import openpyxl as xl

def read_data_from_excel():
	while True:
		try:
			file_path = input("Enter your file path. (Supported formats are: .xlsx, .xlsm, .xltx, and .xltm)\n>>> ")
			wb = xl.load_workbook(file_path) #assuming that the Excel file is located in another folder

			# sheet_name = input("Enter the name of the sheet with your data.\n>>> ")
			# ws = wb[sheet_name] #access the specific sheet
			ws = wb.active
			print(f"Accessing {ws}...")
			break

		except FileNotFoundError as fnfe:
			print(f"{fnfe} file not found. Please check the name or path of the file.")
			continue
	
	while True:
		try:
			#prompt user for the number of data points to access
			num_data_points = int(input("Enter the number of data points to access: "))
			break

		except ValueError:
			print("Invalid input. Please retry.")
			continue

	# #prompt user for the number of data points to access
	# num_data_points = int(input("Enter the number of data points to access: "))


	#SIMPLE PAG IMPORT DATA
	#may mali pala, same data lang ma-store for different data lists amp
	#i'll leave it here lang for future reference or whatever

	# dataset = {} #iterate over the column and store data list in dataset
	# for i in range(1, num_data_points+1):
	# 	data_name = input(f"Enter the name for data {i} that is in your sheet: ")

	# 	data_list = []
	# 	#iterating through the rows available, assuming there are headers
	# 	for row in ws.iter_rows(min_row=2, values_only=True):
	# 		data_list.append(row[0])  #assuming data is in the first column

	# 	dataset[data_name] = data_list

	# #print the dataset
	# for data_name, data_list in dataset.items():
	# 	print(f"{data_name}: {data_list}")



	#ALTERNATIVE MEDJ COMPLEX NA PAG-IMPORT; MAS MAGANDA & MORE FLEXIBILITY BUT PRONE TO ERRORS SO BE CAREFUL

	# dataset = {} #iterate over the column and store data list in dataset
	# for i in range(1, num_data_points+1):
	# 	data_name = input(f"Enter the name for data {i} that is in your sheet: ")

	# 	data_list = []

	# 	for row in ws.iter_rows(): #iterate over all cells in the sheet
	# 		for cell in row: #check if the cell value matches the search value
					
	# 			if cell.value == data_name: #Collect values from cells below the found cell
					
	# 				for below_cell in ws.iter_rows(min_row=cell.row+1, min_col=cell.column, values_only=True):
	# 					if below_cell[0] is not None:
	# 						data_list.append(below_cell[0])

	# 	dataset[data_name] = data_list

	# #print the dataset
	# for data_name, data_list in dataset.items():
	# 	print(f"{data_name}: {data_list}")


	#MAS PINABONGGA PA GD
	dataset = {} #stores the data name as key and the data list as value
	for i in range(1, num_data_points+1): #iterate over the column and store data lists in dataset dict
		while True:
			data_name = input(f"Enter the name for data{i} that is in your sheet: ")
			data_list = []

			for row in ws.iter_rows(): #iterate over all rows in the sheet
				for cell in row: #iterate over all cells in the row
					if cell.value == data_name: #check if the cell value matches the search value
						# cell_address = cell.coordinate #way gamit

						#collect values from cells below the found cell
						for below_cell in ws.iter_rows(min_row=cell.row+1, min_col=cell.column, values_only=True):
							if below_cell[0] is not None:
								data_list.append(below_cell[0])

			if data_list == []: #if data_list with with associated data_name key does not match any value in the sheet
				print(f"No cell has the value '{data_name}'. Please retry.")
				continue
			else:
				dataset[data_name] = data_list
				break #goes to next iteration in outermost for-loop
	
	#print the dataset
	print("Your input summary:")
	for data_name, data_list in dataset.items():
		print(f"\t{data_name}:\t{data_list}")

#START HERE
read_data_from_excel()
